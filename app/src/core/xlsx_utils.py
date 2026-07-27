import re
import io
import time as _time
from typing import Dict, Any, List, Tuple, Optional

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


COLUMN_PATTERNS = {
    "Блок": [
        r"(?i)^блок$", r"(?i)^block$", r"(?i)^група", r"(?i)^group",
        r"(?i)^секція", r"(?i)^section",
    ],
    "Робітник": [
        r"(?i)^робітник$", r"(?i)^worker$", r"(?i)^виконавець",
        r"(?i)^pracownik", r"(?i)^operator", r"(?i)^оператор",
        r"(?i)^ім.?я\s+робітника", r"(?i)^worker\s+name",
    ],
    "Розряд": [
        r"(?i)^розряд$", r"(?i)^rank$", r"(?i)^клас$", r"(?i)^class$",
        r"(?i)^grade$", r"(?i)^рівень", r"(?i)^level",
    ],
    "Обладнання": [
        r"(?i)^обладнання$", r"(?i)^equipment$", r"(?i)^машин", r"(?i)^machine",
        r"(?i)^інструмент", r"(?i)^tool$", r"(?i)^aparat", r"(?i)^станок",
        r"(?i)^типи?\s+обладнан", r"(?i)^equipment\s+type",
    ],
    "№ п/п": [
        r"(?i)^№\s*п/?п", r"(?i)^п/?\s*п", r"(?i)^#?\s*п/?п",
        r"(?i)^seq", r"(?i)^sequence", r"(?i)^порядковий",
        r"(?i)^numbering", r"(?i)^row\s*#?", r"(?i)^№?\s*\d+",
        r"(?i)^інде[кс]", r"(?i)^index$",
    ],
    "№ тех.оп.": [
        r"(?i)^№\s*тех", r"(?i)^tech", r"(?i)^операц",
        r"(?i)^operation\s*(?:number|num|#|no)",
        r"(?i)^tech\s*op", r"(?i)^技术",
    ],
    "Назва технологічної операції": [
        r"(?i)^назва\s+технолог",
        r"(?i)^назва\s+операц", r"(?i)^operation\s*name",
        r"(?i)^name\s+of", r"(?i)^найменування",
        r"(?i)^description", r"(?i)^опис",
        r"(?i)^назва$", r"(?i)^назв",
        r"(?i)^name$",
    ],
    "Затрати часу, хв": [
        r"(?i)^затрати?\s+час",
        r"(?i)^time", r"(?i)^час", r"(?i)^хвил", r"(?i)^min",
        r"(?i)^minutes?", r"(?i)^seconds?", r"(?i)^секунд",
        r"(?i)^витрат", r"(?i)^duration", r"(?i)^тривал",
        r"(?i)^час\s+виконання", r"(?i)^норма\s+часу",
        r"(?i)^трудомісткість", r"(?i)^time\s*cost",
    ],
    "Технічні умови": [
        r"(?i)^технічні\s+умови", r"(?i)^technical",
        r"(?i)^умови$", r"(?i)^conditions?$", r"(?i)^notes?$",
        r"(?i)^примітк", r"(?i)^remark",
    ],
}

TIME_SECONDS_THRESHOLD = 50


def _detect_time_unit(values: List[float]) -> str:
    if not values:
        return "minutes"
    clean = [v for v in values if v > 0]
    if not clean:
        return "minutes"
    avg = sum(clean) / len(clean)
    if avg > TIME_SECONDS_THRESHOLD:
        return "seconds"
    return "minutes"


def _find_column_mapping(headers: List[str]) -> Dict[str, str]:
    mapping = {}
    used = set()
    for target_col, patterns in COLUMN_PATTERNS.items():
        for header in headers:
            if header in used:
                continue
            for pat in patterns:
                if re.search(pat, str(header)):
                    mapping[target_col] = header
                    used.add(header)
                    break
            if target_col in mapping:
                break
    return mapping


def _get_sheet_data(ws) -> Tuple[List[str], List[List[Any]]]:
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []

    header_row_idx = 0
    headers = [str(c) if c is not None else "" for c in all_rows[0]]

    non_empty_count = sum(1 for h in headers if h.strip())
    if non_empty_count < 2 and len(all_rows) > 1:
        header_row_idx = 1
        headers = [str(c) if c is not None else "" for c in all_rows[1]]

    data_rows = []
    for row in all_rows[header_row_idx + 1:]:
        if all(c is None for c in row):
            continue
        data_rows.append(list(row))

    return headers, data_rows


def read_xlsx_to_dataframe(xlsx_bytes: bytes, time_unit: Optional[str] = None) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active

    headers, data_rows = _get_sheet_data(ws)
    wb.close()

    if not headers:
        raise ValueError("XLSX file is empty or has no headers")

    column_mapping = _find_column_mapping(headers)

    time_target = "Затрати часу, хв"
    mapped_time_col = column_mapping.get(time_target)

    raw_time_values = []
    if mapped_time_col:
        time_idx = headers.index(mapped_time_col)
        for row in data_rows:
            if time_idx < len(row):
                val = row[time_idx]
                try:
                    raw_time_values.append(float(str(val).replace(",", ".").replace(" ", "").replace("\xa0", "")))
                except (ValueError, TypeError):
                    pass

    detected_unit = time_unit or _detect_time_unit(raw_time_values)

    renamed = {}
    for target, source in column_mapping.items():
        renamed[source] = target

    df = pd.DataFrame(data_rows, columns=headers)
    df = df.rename(columns=renamed)

    if detected_unit == "seconds" and time_target in df.columns:
        df[time_target] = (
            df[time_target].astype(str)
            .str.replace('\xa0', '', regex=False)
            .str.replace('\u00a0', '', regex=False)
            .str.replace(' ', '', regex=False)
            .str.replace(',', '.', regex=False)
        )
        df[time_target] = pd.to_numeric(df[time_target], errors="coerce").fillna(0.0)
        df[time_target] = (df[time_target] / 60).round(4)

    meta = {
        "column_mapping": column_mapping,
        "detected_time_unit": detected_unit,
        "original_headers": headers,
        "mapped_count": len(column_mapping),
        "total_columns": len(headers),
    }

    return df, meta


def export_to_xlsx(data: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Операції"

    columns = [
        "Блок", "Робітник", "Розряд", "Обладнання",
        "№ п/п", "№ тех.оп.", "Назва технологічної операції",
        "Затрати часу, хв", "Технічні умови",
    ]

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4E48EB", end_color="4E48EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    cell_font = Font(name="Arial", size=10)
    cell_alignment = Alignment(vertical="center")

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row_data in enumerate(data, 2):
        values = [
            row_data.get("block", ""),
            row_data.get("worker", ""),
            row_data.get("rank", ""),
            row_data.get("equipment", ""),
            row_idx - 1,
            row_data.get("techNum", ""),
            row_data.get("name", ""),
            row_data.get("time", 0),
            row_data.get("conditions", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

    col_widths = [12, 18, 8, 18, 8, 12, 35, 16, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()
